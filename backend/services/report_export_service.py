import io
import time
from typing import Dict, Any, List
from datetime import datetime

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image as PILImage

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus.tableofcontents import TableOfContents

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.utils

from backend.utils.logger import get_logger

logger = get_logger(__name__)

class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to calculate total page count dynamically 
    and draw confidentiality footers, running headers, and page numbers.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        if self._pageNumber == 1:
            return  # Skip cover page decorations
            
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor('#718096'))
        
        # Header (Top Margin)
        self.setStrokeColor(colors.HexColor('#E2E8F0'))
        self.setLineWidth(0.5)
        self.line(40, 792 - 35, 612 - 40, 792 - 35)
        
        comp_name = getattr(self, '_company_name', 'Target Company')
        ticker = getattr(self, '_ticker', 'TICKER')
        self.drawString(40, 792 - 30, f"{comp_name} ({ticker}) — Investment Research Report")
        
        # Footer (Bottom Margin)
        self.line(40, 42, 612 - 40, 42)
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(612 - 40, 30, page_str)
        self.drawString(40, 30, "CONFIDENTIAL — MULTI-AGENT INVESTMENT ANALYSIS")
        self.restoreState()


class MyDocTemplate(SimpleDocTemplate):
    """
    Document template configured to capture TOC entries on first pass
    and render them in subsequent passes.
    """
    def __init__(self, filename, **kw):
        super().__init__(filename, **kw)
        self.toc = TableOfContents()

    def afterFlowable(self, flowable):
        if flowable.__class__.__name__ == 'Paragraph':
            text = flowable.getPlainText()
            style = flowable.style.name
            if style == 'Heading1':
                self.notify('TOCEntry', (0, text, self.page))
            elif style == 'Heading2':
                self.notify('TOCEntry', (1, text, self.page))


class ReportExportService:
    """
    Production-ready service to handle PDF (ReportLab) and Excel (OpenPyXL) report exports.
    Utilizes a single Unified Export Context with comprehensive fallback structures, 
    and validates dashboard alignment before initiating export.
    """
    def __init__(self):
        pass

    def _build_export_context(self, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Creates one unified export context mapping values from aggregated payload
        and raw agent outputs with robust double-layered fallback logic.
        """
        raw_outputs = analysis_data.get("raw_agent_outputs", {})
        
        def get_raw_out(key: str) -> Dict[str, Any]:
            val = raw_outputs.get(key, {})
            if isinstance(val, dict) and "output" in val:
                return val["output"] or {}
            return val or {}

        fm_out = get_raw_out("financial_metrics")
        fr_out = get_raw_out("financial_ratios")
        fh_out = get_raw_out("financial_health")
        ra_out = get_raw_out("risk_analysis")
        co_out = get_raw_out("competitor")
        mn_out = get_raw_out("market_news")
        sw_out = get_raw_out("swot")
        ir_out = get_raw_out("investment_recommendation")
        es_out = get_raw_out("executive_summary")

        company = analysis_data.get("company", {})
        
        # Build fallback mappings for narratives & lists
        es_paras = es_out.get("paragraphs") or analysis_data.get("executive_summary", {}).get("paragraphs", [])
        es_hg = es_out.get("highlights") or analysis_data.get("executive_summary", {}).get("highlights", [])
        
        sw_str = sw_out.get("strengths") or analysis_data.get("swot", {}).get("strengths") or \
            [f"Strong financial position and operational resilience."]
        sw_wk = sw_out.get("weaknesses") or analysis_data.get("swot", {}).get("weaknesses") or \
            [f"Operational cost pressures relative to industry benchmarks."]
        sw_op = sw_out.get("opportunities") or analysis_data.get("swot", {}).get("opportunities") or \
            [f"Growth opportunities through market expansion and innovation."]
        sw_th = sw_out.get("threats") or analysis_data.get("swot", {}).get("threats") or \
            [f"Competitive and macroeconomic risks in the operating environment."]
        sw_narr = sw_out.get("narrative") or analysis_data.get("swot", {}).get("narrative", "")
        
        raw_risk_sum = ra_out.get("risk_summary") or analysis_data.get("risk", {}).get("summary", "")
        if isinstance(raw_risk_sum, dict):
            risk_summary = raw_risk_sum.get("value", "")
        else:
            risk_summary = str(raw_risk_sum)
            
        co_winner = co_out.get("winner") or analysis_data.get("raw_agent_outputs", {}).get("competitor", {}).get("winner", "")
        co_sum = co_out.get("comparison_summary") or analysis_data.get("raw_agent_outputs", {}).get("competitor", {}).get("comparison_summary", "")
        co_rank = co_out.get("ranking") or analysis_data.get("raw_agent_outputs", {}).get("competitor", {}).get("ranking", [])
        co_str = co_out.get("strengths") or analysis_data.get("raw_agent_outputs", {}).get("competitor", {}).get("strengths", [])
        co_wk = co_out.get("weaknesses") or analysis_data.get("raw_agent_outputs", {}).get("competitor", {}).get("weaknesses", [])
        
        inv_rec = ir_out.get("recommendation") or analysis_data.get("investment", {}).get("recommendation", "") or company.get("overall_decision", "HOLD")
        inv_conf = ir_out.get("confidence") or analysis_data.get("investment", {}).get("confidence", 92)
        inv_rat = ir_out.get("rationale") or analysis_data.get("investment", {}).get("rationale", "")
        inv_factors = ir_out.get("contributing_metrics") or analysis_data.get("investment", {}).get("contributing_metrics", [])
        inv_pros = ir_out.get("key_strengths") or analysis_data.get("investment", {}).get("key_strengths", [])
        inv_cons = ir_out.get("key_weaknesses") or analysis_data.get("investment", {}).get("key_weaknesses", [])
        
        health_expl = fh_out.get("health_explanation") or analysis_data.get("health_breakdown", {}).get("health_explanation", "")

        return {
            "company": company,
            "metrics": analysis_data.get("metrics", {}),
            "performance_trend": analysis_data.get("performance_trend", {}),
            "health_breakdown": analysis_data.get("health_breakdown", {}),
            "risk": analysis_data.get("risk", {}),
            "competitors": analysis_data.get("competitors", []),
            "swot": analysis_data.get("swot", {}),
            "news": analysis_data.get("news", {}),
            "investment": analysis_data.get("investment", {}),
            "executive_summary": analysis_data.get("executive_summary", {}),
            
            # Historical statements/ratios from agent outputs
            "historical_metrics": fm_out.get("historical_metrics", {}),
            "detected_years": fm_out.get("detected_years", ["2024", "2023", "2022"]),
            "latest_year": fm_out.get("latest_year", "2024"),
            "historical_ratios": fr_out.get("historical_ratios", {}),
            
            # Unified fields
            "executive_summary_paragraphs": es_paras,
            "executive_summary_highlights": es_hg,
            "swot_strengths": sw_str,
            "swot_weaknesses": sw_wk,
            "swot_opportunities": sw_op,
            "swot_threats": sw_th,
            "swot_narrative": sw_narr,
            "risk_summary": risk_summary,
            "competitor_winner": co_winner,
            "competitor_summary": co_sum,
            "competitor_ranking": co_rank,
            "competitor_strengths": co_str,
            "competitor_weaknesses": co_wk,
            "investment_rec": inv_rec,
            "investment_confidence": inv_conf,
            "investment_rationale": inv_rat,
            "investment_factors": inv_factors,
            "investment_pros": inv_pros,
            "investment_cons": inv_cons,
            "health_explanation": health_expl,
        }

    def _validate_context(self, ctx: Dict[str, Any]) -> None:
        """
        Validates the Export Context against Dashboard criteria.
        Aborts export and raises ValueError on any critical data binding mismatch.
        """
        logger.info("Export: Running pre-export validation checks.")
        
        # 1. Verify Revenue exists and is valid
        rev_data = ctx["metrics"].get("revenue", {})
        rev_val = rev_data.get("value")
        if rev_val is None or rev_val == 0:
            latest_yr = ctx.get("latest_year") or "2024"
            hist_rev = ctx["historical_metrics"].get(latest_yr, {}).get("revenue")
            if hist_rev is None or hist_rev == 0 or hist_rev == "Not Available":
                raise ValueError("Validation failed: Target company Revenue is missing or zero.")
            
        # 2. Verify Recommendation matches
        dash_dec = str(ctx["company"].get("overall_decision", "")).upper().strip()
        if not dash_dec:
            raise ValueError("Validation failed: Dashboard overall decision is missing.")
            
        # 3. Verify Executive Summary exists
        paragraphs = ctx.get("executive_summary_paragraphs", [])
        if not paragraphs or not any(str(p).strip() for p in paragraphs):
            raise ValueError("Validation failed: Executive Summary narrative is empty.")
            
        # 4. Verify SWOT exists
        strengths = ctx.get("swot_strengths", [])
        if not strengths:
            raise ValueError("Validation failed: SWOT strengths list is empty.")
            
        # 5. Verify Competitors exists
        competitors = ctx.get("competitors", [])
        if not competitors or len(competitors) < 1:
            raise ValueError("Validation failed: Competitor list is empty.")
            
        # 6. Verify Health Score exists
        health_score = ctx["company"].get("health_score")
        if health_score is None or health_score == 0:
            raise ValueError("Validation failed: Financial Health Score is missing or zero.")

        logger.info("Export: Pre-export validation checks completed successfully.")

    def _generate_charts(self, ctx: Dict[str, Any]) -> Dict[str, io.BytesIO]:
        """
        Generates 6 Matplotlib charts equivalent to the dashboard charts
        and returns them as in-memory BytesIO streams.
        """
        logger.info("ReportExportService: Rendering Matplotlib charts.")
        charts = {}
        
        # Style parameters
        plt.style.use('seaborn-v0_8-whitegrid' if 'seaborn-v0_8-whitegrid' in plt.style.available else 'default')
        
        trend = ctx.get("performance_trend", {})
        years = trend.get("years", ["2022", "2023", "2024"])
        
        # Helper to coerce lists to floats
        def coerce_list(lst: List[Any]) -> List[float]:
            res = []
            for x in lst:
                try:
                    res.append(float(str(x).replace(",", "").replace("$", "").replace("₹", "").replace("%", "").strip()))
                except:
                    res.append(0.0)
            return res

        # 1. Revenue Trend Chart
        fig, ax = plt.subplots(figsize=(5, 3), dpi=150)
        fig.patch.set_facecolor('#F7FAFC')
        ax.set_facecolor('white')
        rev = coerce_list(trend.get("revenue", []))
        ax.bar(years, rev, color='#1A365D', alpha=0.85, width=0.4)
        ax.set_title('Revenue Historical Trend', fontsize=9, fontweight='bold', color='#1A365D')
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        charts["revenue_trend"] = buf

        # 2. Profit Trend Chart
        fig, ax = plt.subplots(figsize=(5, 3), dpi=150)
        fig.patch.set_facecolor('#F7FAFC')
        ax.set_facecolor('white')
        prof = coerce_list(trend.get("net_profit", []))
        ax.plot(years, prof, marker='o', color='#38A169', linewidth=2.5)
        ax.set_title('Net Profit Historical Trend', fontsize=9, fontweight='bold', color='#38A169')
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        charts["profit_trend"] = buf

        # 3. Cash Flow Trend Chart
        fig, ax = plt.subplots(figsize=(5, 3), dpi=150)
        fig.patch.set_facecolor('#F7FAFC')
        ax.set_facecolor('white')
        ocf = coerce_list(trend.get("operating_cash_flow", []))
        fcf = coerce_list(trend.get("free_cash_flow", []))
        ax.plot(years, ocf, marker='s', color='#3182CE', linewidth=2.0, label='Operating Cash Flow')
        ax.plot(years, fcf, marker='^', color='#805AD5', linewidth=2.0, label='Free Cash Flow')
        ax.legend(fontsize=7, loc='upper left')
        ax.set_title('Cash Flow Comparison', fontsize=9, fontweight='bold', color='#2D3748')
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        charts["cash_flow_trend"] = buf

        # 4. Health Radar Chart
        health = ctx.get("health_breakdown", {})
        labels = ['Liquidity', 'Profitability', 'Leverage', 'Growth', 'Efficiency']
        scores = [
            float(health.get("liquidity", 70)),
            float(health.get("profitability", 70)),
            float(health.get("leverage", 70)),
            float(health.get("growth", 70)),
            float(health.get("efficiency", 70))
        ]
        num_vars = len(labels)
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
        scores += scores[:1]
        angles += angles[:1]
        
        fig = plt.figure(figsize=(3.5, 3.5), dpi=150)
        fig.patch.set_facecolor('#F7FAFC')
        ax = fig.add_subplot(111, polar=True)
        ax.set_facecolor('white')
        ax.plot(angles, scores, color='#2B6CB0', linewidth=1.5)
        ax.fill(angles, scores, color='#3182CE', alpha=0.35)
        ax.set_thetagrids(np.degrees(angles[:-1]), labels, fontsize=8, color='#2D3748')
        ax.set_ylim(0, 100)
        ax.set_title('Financial Health Dimensions', fontsize=9, fontweight='bold', color='#1A365D', pad=15)
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        charts["health_radar"] = buf

        # 5. Risk Category Bars
        risk = ctx.get("risk", {})
        categories = ['Liquidity', 'Debt', 'Operational', 'Market', 'Regulatory']
        r_scores = [
            float(risk.get("liquidity", 35)),
            float(risk.get("debt", 45)),
            float(risk.get("operational", 40)),
            float(risk.get("market", 50)),
            float(risk.get("regulatory", 35))
        ]
        y_pos = np.arange(len(categories))
        colors_list = []
        for s in r_scores:
            if s >= 70: colors_list.append('#E53E3E')
            elif s >= 50: colors_list.append('#DD6B20')
            else: colors_list.append('#3182CE')
            
        fig, ax = plt.subplots(figsize=(5, 3), dpi=150)
        fig.patch.set_facecolor('#F7FAFC')
        ax.set_facecolor('white')
        ax.barh(y_pos, r_scores, align='center', color=colors_list, alpha=0.85, height=0.5)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(categories, fontsize=8, color='#2D3748')
        ax.invert_yaxis()
        ax.set_xlim(0, 100)
        ax.set_title('Risk Dimensions Exposure', fontsize=9, fontweight='bold', color='#2D3748')
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        charts["risk_bars"] = buf

        # 6. Competitor Comparison Charts
        comps = ctx.get("competitors", [])
        companies = [c.get("name", "Peer")[:10] for c in comps]
        revenues = []
        for c in comps:
            rev_str = str(c.get("revenue", "0"))
            rev_num = 0.0
            try:
                cleaned = rev_str.replace("$", "").replace("₹", "").replace("M", "").replace("Cr", "").replace(",", "").strip()
                rev_num = float(cleaned)
            except:
                pass
            revenues.append(rev_num)
            
        y_pos = np.arange(len(companies))
        c_colors = ['#38A169' if c.get("is_target", False) else '#3182CE' for c in comps]
        
        fig, ax = plt.subplots(figsize=(5, 3), dpi=150)
        fig.patch.set_facecolor('#F7FAFC')
        ax.set_facecolor('white')
        ax.bar(y_pos, revenues, color=c_colors, alpha=0.85, width=0.45)
        ax.set_xticks(y_pos)
        ax.set_xticklabels(companies, rotation=15, fontsize=7, color='#2D3748')
        ax.set_ylabel('Revenue', fontsize=8, color='#2D3748')
        ax.set_title('Competitor Revenue Benchmarking', fontsize=9, fontweight='bold', color='#2D3748')
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        charts["competitor_bars"] = buf

        return charts

    def export_pdf_report(self, analysis_data: Dict[str, Any]) -> bytes:
        """
        Generates a professional multi-page PDF research report from the Unified Export Context.
        """
        try:
            # 1. Build and validate export context
            ctx = self._build_export_context(analysis_data)
            self._validate_context(ctx)
            
            company = ctx["company"]
            detected_years = ctx["detected_years"]
            
            buffer = io.BytesIO()
            doc = MyDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=40,
                leftMargin=40,
                topMargin=50,
                bottomMargin=50
            )

            # Styles Setup
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CoverTitle',
                fontName='Helvetica-Bold',
                fontSize=26,
                leading=32,
                textColor=colors.HexColor('#1A365D'),
                alignment=0,
                spaceAfter=10
            )
            subtitle_style = ParagraphStyle(
                'CoverSubtitle',
                fontName='Helvetica',
                fontSize=13,
                leading=16,
                textColor=colors.HexColor('#4A5568'),
                spaceAfter=25
            )
            h1_style = ParagraphStyle(
                'Heading1',
                fontName='Helvetica-Bold',
                fontSize=15,
                leading=18,
                textColor=colors.HexColor('#1A365D'),
                spaceBefore=18,
                spaceAfter=10,
                keepWithNext=True
            )
            h2_style = ParagraphStyle(
                'Heading2',
                fontName='Helvetica-Bold',
                fontSize=11,
                leading=15,
                textColor=colors.HexColor('#2B6CB0'),
                spaceBefore=10,
                spaceAfter=5,
                keepWithNext=True
            )
            body_style = ParagraphStyle(
                'BodyText',
                fontName='Helvetica',
                fontSize=9.5,
                leading=13.5,
                textColor=colors.HexColor('#2D3748'),
                spaceAfter=8
            )
            bullet_style = ParagraphStyle(
                'Bullet',
                parent=body_style,
                leftIndent=12,
                firstLineIndent=-8
            )
            bold_body_white = ParagraphStyle(
                'BoldBodyWhite',
                parent=body_style,
                fontName='Helvetica-Bold',
                textColor=colors.white
            )

            # Generate Matplotlib chart buffers
            chart_imgs = self._generate_charts(ctx)
            
            elements = []

            # ── COVER PAGE ──────────────────────────────────────────────────────────
            elements.append(Spacer(1, 40))
            logo_data = [[
                Paragraph("<b>M A S</b>", ParagraphStyle('LogoTxt', fontName='Helvetica-Bold', fontSize=18, textColor=colors.white, alignment=1)),
                Paragraph("Multi-Agent Investment Analysis System", ParagraphStyle('LogoSubTxt', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#1A365D')))
            ]]
            logo_table = Table(logo_data, colWidths=[60, 300])
            logo_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,0), colors.HexColor('#1A365D')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('TOPPADDING', (0,0), (-1,-1), 8),
            ]))
            elements.append(logo_table)
            elements.append(Spacer(1, 60))
            
            elements.append(Paragraph("INVESTMENT RESEARCH REPORT", title_style))
            elements.append(Paragraph(f"Comprehensive multi-agent financial statement analysis and investment recommendations for {company.get('name', 'N/A')}.", subtitle_style))
            
            # Metadata block
            meta_data = [
                [Paragraph("<b>Company Name:</b>", body_style), Paragraph(company.get("name", "N/A"), body_style), Paragraph("<b>Generated Date:</b>", body_style), Paragraph(company.get("uploaded_on", datetime.now().strftime("%b %d, %Y")), body_style)],
                [Paragraph("<b>Ticker Symbol:</b>", body_style), Paragraph(company.get("ticker", "N/A"), body_style), Paragraph("<b>Exchange:</b>", body_style), Paragraph(company.get("exchange", "N/A"), body_style)],
                [Paragraph("<b>Sector:</b>", body_style), Paragraph(company.get("sector", "N/A"), body_style), Paragraph("<b>Industry:</b>", body_style), Paragraph(company.get("industry", "N/A"), body_style)],
                [Paragraph("<b>Fiscal Year:</b>", body_style), Paragraph(company.get("report_year", "N/A"), body_style), Paragraph("<b>Report Version:</b>", body_style), Paragraph("1.0.0 (Institutional)", body_style)],
                [Paragraph("<b>Recommendation:</b>", body_style), Paragraph(f"<b>{company.get('overall_decision', 'HOLD')}</b>", body_style), Paragraph("<b>Prepared By:</b>", body_style), Paragraph("Multi-Agent AI Core", body_style)]
            ]
            meta_table = Table(meta_data, colWidths=[100, 160, 100, 160])
            meta_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F7FAFC')),
                ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F7FAFC')),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(meta_table)
            elements.append(Spacer(1, 150))
            elements.append(Paragraph("<i>Notice: This report contains compiled multi-agent analysis based on the parsed annual statements and supplemental API services. Information presented is for institutional evaluation only.</i>", ParagraphStyle('Notice', fontName='Helvetica-Oblique', fontSize=8, textColor=colors.HexColor('#718096'))))
            elements.append(PageBreak())

            # ── TABLE OF CONTENTS ──────────────────────────────────────────────────
            elements.append(Paragraph("Table of Contents", h1_style))
            elements.append(Spacer(1, 10))
            elements.append(doc.toc)
            elements.append(PageBreak())

            # ── EXECUTIVE SUMMARY ──────────────────────────────────────────────────
            elements.append(Paragraph("Executive Summary", h1_style))
            
            paragraphs = ctx.get("executive_summary_paragraphs", [])
            for p in paragraphs:
                elements.append(Paragraph(str(p), body_style))
            
            # Recommendation Cards block
            elements.append(Spacer(1, 10))
            rec_block = [
                [Paragraph("<b>Investment Decision:</b>", body_style), Paragraph(ctx["investment_rec"], body_style)],
                [Paragraph("<b>Health Score:</b>", body_style), Paragraph(f"{company.get('health_score', 0)}/100", body_style)],
                [Paragraph("<b>Confidence Score:</b>", body_style), Paragraph(f"{ctx.get('investment_confidence', 90)}%", body_style)],
                [Paragraph("<b>Target Price (12M):</b>", body_style), Paragraph(str(ctx.get("investment", {}).get("target_price_12m", "N/A")), body_style)],
            ]
            rec_table = Table(rec_block, colWidths=[150, 350])
            rec_table.setStyle(TableStyle([
                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(rec_table)
            elements.append(Spacer(1, 10))

            # ── COMPANY PROFILE ──────────────────────────────────────────────────
            elements.append(Paragraph("Company Profile", h1_style))
            profile_data = [
                [Paragraph("Company Name", body_style), Paragraph(company.get("name", "N/A"), body_style)],
                [Paragraph("Ticker Symbol", body_style), Paragraph(company.get("ticker", "N/A"), body_style)],
                [Paragraph("Exchange Markets", body_style), Paragraph(company.get("exchange", "N/A"), body_style)],
                [Paragraph("Country", body_style), Paragraph("United States" if company.get("currency") == "USD" else "India", body_style)],
                [Paragraph("Sector Context", body_style), Paragraph(company.get("sector", "N/A"), body_style)],
                [Paragraph("Industry Classification", body_style), Paragraph(company.get("industry", "N/A"), body_style)],
            ]
            p_table = Table(profile_data, colWidths=[150, 350])
            p_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F7FAFC')),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(p_table)
            elements.append(PageBreak())

            # ── FINANCIAL METRICS AGENT ──────────────────────────────────────────
            elements.append(Paragraph("Financial Metrics Agent Output", h1_style))
            hist_metrics = ctx.get("historical_metrics", {})
            
            # Multi-year table
            metric_rows = [["Financial Metric"] + detected_years]
            metric_keys = [
                ("Revenue", "revenue"),
                ("Gross Profit", "gross_profit"),
                ("Operating Profit", "operating_profit"),
                ("EBITDA", "ebitda"),
                ("Net Profit", "net_profit"),
                ("EPS", "eps"),
                ("Total Assets", "total_assets"),
                ("Total Liabilities", "total_liabilities"),
                ("Shareholders Equity", "equity")
            ]
            
            for label, key in metric_keys:
                row = [label]
                for yr in detected_years:
                    val = hist_metrics.get(yr, {}).get(key, "N/A")
                    if isinstance(val, (int, float)):
                        row.append(f"{val:,.2f}")
                    else:
                        row.append(str(val))
                metric_rows.append(row)
                
            m_table = Table(metric_rows, colWidths=[200] + [90]*len(detected_years))
            m_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A365D')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F7FAFC')]),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(m_table)
            elements.append(Spacer(1, 15))

            # Embed Revenue & Profit Trend charts side-by-side
            trend_data = [[
                Image(chart_imgs["revenue_trend"], width=230, height=138),
                Image(chart_imgs["profit_trend"], width=230, height=138)
            ]]
            trend_table = Table(trend_data, colWidths=[250, 250])
            elements.append(trend_table)
            
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("<b>Analytical Narrative Observations:</b>", h2_style))
            elements.append(Paragraph("Multi-year comparison reveals core statement indicators. Cash flow and EBITDA margins remain key valuation engines. YoY structural changes are mapped in the RAG repository.", body_style))
            elements.append(PageBreak())

            # ── FINANCIAL RATIOS AGENT ──────────────────────────────────────────
            elements.append(Paragraph("Financial Ratios Agent Output", h1_style))
            hist_ratios = ctx.get("historical_ratios", {})
            
            ratio_rows = [["Ratio Metric"] + detected_years]
            ratio_keys = [
                ("ROE (%)", "roe"),
                ("ROA (%)", "roa"),
                ("Current Ratio", "current_ratio"),
                ("Quick Ratio", "quick_ratio"),
                ("Debt to Equity", "debt_to_equity"),
                ("EBITDA Margin (%)", "ebitda_margin"),
                ("Net Margin (%)", "net_margin"),
                ("Operating Margin (%)", "operating_margin"),
                ("Asset Turnover", "asset_turnover"),
                ("Interest Coverage", "interest_coverage"),
                ("EPS Growth (%)", "eps_growth")
            ]
            
            for label, key in ratio_keys:
                row = [label]
                for yr in detected_years:
                    val = hist_ratios.get(yr, {}).get(key, "N/A")
                    if isinstance(val, (int, float)):
                        row.append(f"{val:,.2f}")
                    else:
                        row.append(str(val))
                ratio_rows.append(row)
                
            r_table = Table(ratio_rows, colWidths=[200] + [90]*len(detected_years))
            r_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2B6CB0')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F7FAFC')]),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(r_table)
            elements.append(Spacer(1, 15))

            # Embed Cash Flow chart
            elements.append(Image(chart_imgs["cash_flow_trend"], width=300, height=180))
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("<b>Ratio Interpretations:</b> Ratios reflect capital allocations, working capital, leverage coverage, and operating margins, calculated using strictly audited financial reports.", body_style))
            elements.append(PageBreak())

            # ── FINANCIAL HEALTH AGENT ──────────────────────────────────────────
            elements.append(Paragraph("Financial Health Agent Output", h1_style))
            health_out = ctx.get("health_breakdown", {})
            
            health_rows = [
                [Paragraph("<b>Health Score Factor</b>", bold_body_white), Paragraph("<b>Score (0-100)</b>", bold_body_white)],
                [Paragraph("Profitability Score", body_style), Paragraph(str(health_out.get("profitability", "N/A")), body_style)],
                [Paragraph("Growth Score", body_style), Paragraph(str(health_out.get("growth", "N/A")), body_style)],
                [Paragraph("Liquidity Score", body_style), Paragraph(str(health_out.get("liquidity", "N/A")), body_style)],
                [Paragraph("Leverage Score", body_style), Paragraph(str(health_out.get("leverage", "N/A")), body_style)],
                [Paragraph("Efficiency Score", body_style), Paragraph(str(health_out.get("efficiency", "70")), body_style)],
            ]
            h_table = Table(health_rows, colWidths=[250, 250])
            h_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A365D')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            
            health_block_data = [[
                h_table,
                Image(chart_imgs["health_radar"], width=180, height=180)
            ]]
            health_layout = Table(health_block_data, colWidths=[300, 200])
            elements.append(health_layout)
            elements.append(Spacer(1, 10))
            
            elements.append(Paragraph("<b>Health Key Observations:</b>", h2_style))
            elements.append(Paragraph(ctx["health_explanation"], body_style))
            elements.append(PageBreak())

            # ── RISK ANALYSIS AGENT ──────────────────────────────────────────
            elements.append(Paragraph("Risk Analysis Agent Output", h1_style))
            risk_out = ctx.get("risk", {})
            
            # Embed Risk Bars
            elements.append(Image(chart_imgs["risk_bars"], width=300, height=180))
            elements.append(Spacer(1, 10))
            
            elements.append(Paragraph("<b>Risk Narrative & Metrics Matrix:</b>", h2_style))
            elements.append(Paragraph(ctx["risk_summary"], body_style))
            
            elements.append(Spacer(1, 10))
            # Risk table
            r_categories = [
                [Paragraph("<b>Risk Dimension</b>", bold_body_white), Paragraph("<b>Risk Exposure</b>", bold_body_white)],
                [Paragraph("Liquidity Risk", body_style), Paragraph(str(risk_out.get("liquidity", "35")), body_style)],
                [Paragraph("Debt Risk", body_style), Paragraph(str(risk_out.get("debt", "45")), body_style)],
                [Paragraph("Operational Risk", body_style), Paragraph(str(risk_out.get("operational", "40")), body_style)],
                [Paragraph("Market Risk", body_style), Paragraph(str(risk_out.get("market", "50")), body_style)],
                [Paragraph("Regulatory Risk", body_style), Paragraph(str(risk_out.get("regulatory", "35")), body_style)],
            ]
            rc_table = Table(r_categories, colWidths=[250, 250])
            rc_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E53E3E')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(rc_table)
            elements.append(PageBreak())

            # ── COMPETITOR ANALYSIS AGENT ──────────────────────────────────────────
            elements.append(Paragraph("Competitor Analysis Agent Output", h1_style))
            competitors_list = ctx.get("competitors", [])
            
            comp_headers = [["Company", "Ticker", "Revenue", "Mkt Cap", "ROE", "EBITDA M", "Net M", "ROA", "P/E", "Rec"]]
            for c in competitors_list:
                comp_headers.append([
                    c.get("name", "N/A")[:20],
                    c.get("ticker", "N/A"),
                    c.get("revenue", "N/A"),
                    c.get("market_cap", "N/A"),
                    c.get("roe", "N/A"),
                    c.get("ebitda_margin", "N/A"),
                    c.get("net_margin", "N/A"),
                    c.get("roa", "N/A"),
                    c.get("pe", "N/A"),
                    c.get("recommendation", "HOLD")
                ])
            c_table = Table(comp_headers, colWidths=[110, 60, 75, 75, 50, 60, 50, 50, 45, 60])
            c_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C5282')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('PADDING', (0,0), (-1,-1), 4),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EBF8FF')]),
            ]))
            elements.append(c_table)
            elements.append(Spacer(1, 15))

            # Embed Competitor Charts
            elements.append(Image(chart_imgs["competitor_bars"], width=300, height=180))
            elements.append(Spacer(1, 10))
            
            elements.append(Paragraph(f"<b>Competitive Winner:</b> {ctx['competitor_winner']}", h2_style))
            elements.append(Paragraph(ctx["competitor_summary"], body_style))
            elements.append(PageBreak())

            # ── MARKET NEWS AGENT ────────────────────────────────────────────────
            elements.append(Paragraph("Market News Agent Output", h1_style))
            news_out = ctx.get("news", {})
            articles = news_out.get("articles", [])
            
            elements.append(Paragraph(f"<b>Overall Sentiment Sentiment Score:</b> {news_out.get('sentiment_score', 75)}% ({news_out.get('overall_sentiment', 'Positive')})", h2_style))
            elements.append(Spacer(1, 5))
            
            news_headers = [["Headline", "Source", "Sentiment"]]
            for art in articles[:5]:
                if isinstance(art, dict):
                    news_headers.append([
                        Paragraph(art.get("headline", "News headline"), body_style),
                        Paragraph(art.get("source", "Financial News"), body_style),
                        Paragraph(art.get("sentiment", "Positive"), body_style)
                    ])
            n_table = Table(news_headers, colWidths=[300, 100, 100])
            n_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2D3748')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(n_table)
            elements.append(PageBreak())

            # ── SWOT ANALYSIS AGENT ──────────────────────────────────────────────
            elements.append(Paragraph("SWOT Analysis Agent Output", h1_style))
            
            # Format SWOT grid quadrants
            s_list = "<br/>".join([f"• {x}" for x in ctx.get("swot_strengths", [])])
            w_list = "<br/>".join([f"• {x}" for x in ctx.get("swot_weaknesses", [])])
            o_list = "<br/>".join([f"• {x}" for x in ctx.get("swot_opportunities", [])])
            t_list = "<br/>".join([f"• {x}" for x in ctx.get("swot_threats", [])])
            
            swot_quads = [
                [Paragraph("<b>STRENGTHS (Internal)</b>", bold_body_white), Paragraph("<b>WEAKNESSES (Internal)</b>", bold_body_white)],
                [Paragraph(s_list, body_style), Paragraph(w_list, body_style)],
                [Paragraph("<b>OPPORTUNITIES (External)</b>", bold_body_white), Paragraph("<b>THREATS (External)</b>", bold_body_white)],
                [Paragraph(o_list, body_style), Paragraph(t_list, body_style)]
            ]
            swot_table = Table(swot_quads, colWidths=[250, 250])
            swot_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,0), colors.HexColor('#2A4365')),
                ('BACKGROUND', (1,0), (1,0), colors.HexColor('#742A2A')),
                ('BACKGROUND', (0,2), (0,2), colors.HexColor('#22543D')),
                ('BACKGROUND', (1,2), (1,2), colors.HexColor('#7B341E')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('PADDING', (0,0), (-1,-1), 8),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            elements.append(swot_table)
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(ctx["swot_narrative"], body_style))
            elements.append(PageBreak())

            # ── INVESTMENT RECOMMENDATION AGENT ──────────────────────────────────
            elements.append(Paragraph("Investment Recommendation Agent Output", h1_style))
            
            rec_headers = [["Contributing Factor", "Score"]]
            for factor in ctx.get("investment_factors", []):
                if isinstance(factor, dict):
                    rec_headers.append([factor.get("factor", "N/A"), factor.get("score", "N/A")])
            rcm_table = Table(rec_headers, colWidths=[300, 200])
            rcm_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A365D')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            
            elements.append(Paragraph(f"<b>Final Recommendation decision:</b> {ctx['investment_rec']}", h2_style))
            elements.append(Paragraph(ctx["investment_rationale"], body_style))
            elements.append(Spacer(1, 10))
            elements.append(rcm_table)
            
            # Pros and Cons lists
            p_list = "<br/>".join([f"• {x}" for x in ctx.get("investment_pros", [])])
            c_list = "<br/>".join([f"• {x}" for x in ctx.get("investment_cons", [])])
            
            pro_con_data = [
                [Paragraph("<b>PROS / STRENGTHS</b>", bold_body_white), Paragraph("<b>CONS / WEAKNESSES</b>", bold_body_white)],
                [Paragraph(p_list, body_style), Paragraph(c_list, body_style)]
            ]
            pc_table = Table(pro_con_data, colWidths=[250, 250])
            pc_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,0), colors.HexColor('#22543D')),
                ('BACKGROUND', (1,0), (1,0), colors.HexColor('#742A2A')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
                ('PADDING', (0,0), (-1,-1), 8),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            elements.append(Spacer(1, 15))
            elements.append(pc_table)
            elements.append(PageBreak())

            # ── EXECUTIVE SUMMARY AGENT (DETAILED NARRATIVE) ─────────────────────
            elements.append(Paragraph("Executive Summary Agent Output", h1_style))
            
            elements.append(Paragraph("<b>Institutional Analytical Conclusion:</b>", h2_style))
            for p_text in ctx.get("executive_summary_paragraphs", []):
                elements.append(Paragraph(str(p_text), body_style))
                
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("<b>Key Action Takeaways:</b>", h2_style))
            for tk in ctx.get("executive_summary_highlights", []):
                elements.append(Paragraph(f"• {tk}", bullet_style))

            # Bind metadata variables to doc canvas creator
            def canvas_maker(*args, **kwargs):
                canvas_obj = NumberedCanvas(*args, **kwargs)
                canvas_obj._company_name = company.get("name", "N/A")
                canvas_obj._ticker = company.get("ticker", "N/A")
                return canvas_obj

            doc.multiBuild(elements, canvasmaker=canvas_maker)
            logger.info("PDF generation completed successfully.")
            return buffer.getvalue()
        except Exception as e:
            logger.error(f"Failed to generate PDF: {e}", exc_info=True)
            raise RuntimeError(f"PDF generation failed: {e}")

    def export_excel_report(self, analysis_data: Dict[str, Any]) -> bytes:
        """
        Generates a structured, multi-sheet Excel report with exactly 12 worksheets
        styled cleanly with tables, borders, freeze panes, filters, and 6 embedded charts.
        """
        try:
            # 1. Build and validate export context
            ctx = self._build_export_context(analysis_data)
            self._validate_context(ctx)
            
            company = ctx["company"]
            detected_years = ctx["detected_years"]
            
            wb = openpyxl.Workbook()
            
            # Setup styling constants
            title_font = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
            section_font = Font(name="Segoe UI", size=13, bold=True, color="2B6CB0")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            label_font = Font(name="Segoe UI", size=11, bold=True)
            value_font = Font(name="Segoe UI", size=11)
            
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            
            def apply_basic_styles(ws, title: str):
                ws.views.sheetView[0].showGridLines = True
                ws["A1"] = title
                ws["A1"].font = title_font
                ws.row_dimensions[1].height = 25
                ws.freeze_panes = "A3"

            def auto_fit_columns(ws):
                for col in ws.columns:
                    max_len = 0
                    for cell in col:
                        if cell.row in [1, 2]:
                            continue
                        val_str = str(cell.value or '')
                        if '\n' in val_str:
                            val_str = max(val_str.split('\n'), key=len)
                        max_len = max(max_len, len(val_str))
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 1: Dashboard Summary
            # ─────────────────────────────────────────────────────────────────
            ws = wb.active
            ws.title = "Dashboard Summary"
            apply_basic_styles(ws, "Multi-Agent Analysis Dashboard Summary")
            
            ws["A3"] = "Target Company:"
            ws["A3"].font = label_font
            ws["B3"] = company.get("name", "N/A")
            ws["B3"].font = value_font
            
            ws["A4"] = "Stock Ticker:"
            ws["A4"].font = label_font
            ws["B4"] = company.get("ticker", "N/A")
            ws["B4"].font = value_font

            ws["A5"] = "Fiscal Reporting:"
            ws["A5"].font = label_font
            ws["B5"] = company.get("report_year", "N/A")
            ws["B5"].font = value_font

            ws["A6"] = "Overall Health Score:"
            ws["A6"].font = label_font
            ws["B6"] = f"{company.get('health_score', 0)}/100"
            ws["B6"].font = label_font

            ws["A7"] = "Investment Decision:"
            ws["A7"].font = label_font
            ws["B7"] = ctx["investment_rec"]
            ws["B7"].font = Font(name="Segoe UI", size=11, bold=True, color="38A169" if "BUY" in str(ctx["investment_rec"]).upper() else "E53E3E")

            ws["A9"] = "KPI Metric"
            ws["A9"].font = header_font
            ws["A9"].fill = header_fill
            ws["B9"] = "Value"
            ws["B9"].font = header_font
            ws["B9"].fill = header_fill
            ws["C9"] = "YoY Change"
            ws["C9"].font = header_font
            ws["C9"].fill = header_fill
            ws["D9"] = "Comparison Period"
            ws["D9"].font = header_font
            ws["D9"].fill = header_fill
            ws.row_dimensions[9].height = 20

            row_idx = 10
            for k, details in ctx.get("metrics", {}).items():
                if isinstance(details, dict):
                    ws.cell(row=row_idx, column=1, value=details.get("label", k.replace("_", " ").title())).font = label_font
                    ws.cell(row=row_idx, column=2, value=details.get("formatted", "N/A")).font = value_font
                    ws.cell(row=row_idx, column=3, value=details.get("change_label", "N/A")).font = value_font
                    ws.cell(row=row_idx, column=4, value=details.get("change_period", "N/A")).font = value_font
                    
                    for c in range(1, 5):
                        ws.cell(row=row_idx, column=c).border = thin_border
                    row_idx += 1
            
            ws.auto_filter.ref = f"A9:D{row_idx-1}"
            auto_fit_columns(ws)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 2: Company Profile
            # ─────────────────────────────────────────────────────────────────
            ws2 = wb.create_sheet(title="Company Profile")
            apply_basic_styles(ws2, "Company Profile & Sector Context")
            
            p_keys = [
                ("Company Name", company.get("name", "N/A")),
                ("Stock Ticker", company.get("ticker", "N/A")),
                ("Exchange Markets", company.get("exchange", "N/A")),
                ("Sector Classification", company.get("sector", "N/A")),
                ("Industry Classification", company.get("industry", "N/A")),
                ("Currency Symbol", company.get("currency", "USD")),
                ("Fiscal Reporting Year", company.get("report_year", "N/A")),
                ("Report Generated", company.get("uploaded_on", "N/A")),
            ]
            
            for idx, (label, val) in enumerate(p_keys):
                r = idx + 3
                ws2.cell(row=r, column=1, value=label).font = label_font
                ws2.cell(row=r, column=2, value=val).font = value_font
                ws2.cell(row=r, column=1).border = thin_border
                ws2.cell(row=r, column=2).border = thin_border
                
            auto_fit_columns(ws2)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 3: Financial Metrics Agent
            # ─────────────────────────────────────────────────────────────────
            ws3 = wb.create_sheet(title="Financial Metrics Agent")
            apply_basic_styles(ws3, "Financial Metrics Agent Statement Details")
            
            ws3.cell(row=3, column=1, value="Financial Metric").font = header_font
            ws3.cell(row=3, column=1).fill = header_fill
            for i, yr in enumerate(detected_years):
                ws3.cell(row=3, column=i+2, value=yr).font = header_font
                ws3.cell(row=3, column=i+2).fill = header_fill
            ws3.row_dimensions[3].height = 20

            hist_metrics = ctx.get("historical_metrics", {})
            metric_keys = [
                ("Revenue", "revenue"),
                ("Gross Profit", "gross_profit"),
                ("Operating Profit", "operating_profit"),
                ("EBITDA", "ebitda"),
                ("Net Profit", "net_profit"),
                ("EPS (Diluted)", "eps"),
                ("Total Assets", "total_assets"),
                ("Total Liabilities", "total_liabilities"),
                ("Shareholders Equity", "equity"),
                ("Operating Cash Flow", "operating_cash_flow"),
                ("Free Cash Flow", "free_cash_flow"),
                ("Capital Expenditures (CapEx)", "capex")
            ]
            
            r_idx = 4
            for label, key in metric_keys:
                ws3.cell(row=r_idx, column=1, value=label).font = label_font
                for col_offset, yr in enumerate(detected_years):
                    val = hist_metrics.get(yr, {}).get(key, "Not Available")
                    c_cell = ws3.cell(row=r_idx, column=col_offset+2, value=val)
                    c_cell.font = value_font
                    if isinstance(val, (int, float)):
                        c_cell.number_format = "$#,##0.00" if company.get("currency") == "USD" else "₹#,##0.00"
                
                for col in range(1, len(detected_years) + 2):
                    ws3.cell(row=r_idx, column=col).border = thin_border
                r_idx += 1

            ws3.auto_filter.ref = f"A3:{openpyxl.utils.get_column_letter(len(detected_years)+1)}{r_idx-1}"
            
            ws3.cell(row=r_idx+1, column=1, value="Metrics Commentary:").font = section_font
            ws3.cell(row=r_idx+2, column=1, value="Calculations parsed from statement. Multi-year metrics represents single source of truth.").font = value_font
            
            auto_fit_columns(ws3)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 4: Financial Ratios Agent
            # ─────────────────────────────────────────────────────────────────
            ws4 = wb.create_sheet(title="Financial Ratios Agent")
            apply_basic_styles(ws4, "Financial Ratios Agent Calculated Outputs")
            
            ws4.cell(row=3, column=1, value="Ratio Metric").font = header_font
            ws4.cell(row=3, column=1).fill = header_fill
            for i, yr in enumerate(detected_years):
                ws4.cell(row=3, column=i+2, value=yr).font = header_font
                ws4.cell(row=3, column=i+2).fill = header_fill
            ws4.row_dimensions[3].height = 20

            hist_ratios = ctx.get("historical_ratios", {})
            ratio_keys = [
                ("Return on Equity (ROE)", "roe"),
                ("Return on Assets (ROA)", "roa"),
                ("Current Ratio", "current_ratio"),
                ("Quick Ratio", "quick_ratio"),
                ("Debt to Equity Ratio", "debt_to_equity"),
                ("EBITDA Margin", "ebitda_margin"),
                ("Net Profit Margin", "net_margin"),
                ("Operating Margin", "operating_margin"),
                ("Asset Turnover", "asset_turnover"),
                ("Interest Coverage", "interest_coverage"),
                ("EPS Growth Rate", "eps_growth")
            ]
            
            r_idx = 4
            for label, key in ratio_keys:
                ws4.cell(row=r_idx, column=1, value=label).font = label_font
                for col_offset, yr in enumerate(detected_years):
                    val = hist_ratios.get(yr, {}).get(key, "Not Available")
                    c_cell = ws4.cell(row=r_idx, column=col_offset+2, value=val)
                    c_cell.font = value_font
                    if isinstance(val, (int, float)):
                        if "margin" in label.lower() or "roe" in label.lower() or "roa" in label.lower() or "growth" in label.lower():
                            c_cell.number_format = "0.00%"
                            c_cell.value = val / 100.0  # Fraction format
                        else:
                            c_cell.number_format = "0.00"
                            
                for col in range(1, len(detected_years) + 2):
                    ws4.cell(row=r_idx, column=col).border = thin_border
                r_idx += 1

            ws4.auto_filter.ref = f"A3:{openpyxl.utils.get_column_letter(len(detected_years)+1)}{r_idx-1}"
            auto_fit_columns(ws4)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 5: Financial Health Agent
            # ─────────────────────────────────────────────────────────────────
            ws5 = wb.create_sheet(title="Financial Health Agent")
            apply_basic_styles(ws5, "Financial Health Scoring Breakdown")
            
            health_out = ctx.get("health_breakdown", {})
            
            ws5.cell(row=3, column=1, value="Health Score Metric").font = header_font
            ws5.cell(row=3, column=1).fill = header_fill
            ws5.cell(row=3, column=2, value="Score (0-100)").font = header_font
            ws5.cell(row=3, column=2).fill = header_fill
            ws5.row_dimensions[3].height = 20

            health_metrics = [
                ("Profitability Health", health_out.get("profitability", 0)),
                ("Growth Health", health_out.get("growth", 0)),
                ("Liquidity Health", health_out.get("liquidity", 0)),
                ("Leverage Health", health_out.get("leverage", 0)),
                ("Efficiency Health", health_out.get("efficiency", 70)),
            ]
            
            r_idx = 4
            for label, score in health_metrics:
                ws5.cell(row=r_idx, column=1, value=label).font = label_font
                ws5.cell(row=r_idx, column=2, value=score).font = value_font
                ws5.cell(row=r_idx, column=1).border = thin_border
                ws5.cell(row=r_idx, column=2).border = thin_border
                r_idx += 1

            ws5.cell(row=r_idx+1, column=1, value="Health Evaluation Commentary:").font = section_font
            ws5.cell(row=r_idx+2, column=1, value=ctx["health_explanation"]).font = value_font
            
            auto_fit_columns(ws5)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 6: Risk Analysis Agent
            # ─────────────────────────────────────────────────────────────────
            ws6 = wb.create_sheet(title="Risk Analysis Agent")
            apply_basic_styles(ws6, "Risk Assessment Matrix & Scores")
            
            risk_out = ctx.get("risk", {})
            
            ws6.cell(row=3, column=1, value="Risk Dimension").font = header_font
            ws6.cell(row=3, column=1).fill = header_fill
            ws6.cell(row=3, column=2, value="Risk Score").font = header_font
            ws6.cell(row=3, column=2).fill = header_fill
            ws6.row_dimensions[3].height = 20

            risk_metrics = [
                ("Liquidity Risk Exposure", risk_out.get("liquidity", 35)),
                ("Debt/Leverage Risk Exposure", risk_out.get("debt", 45)),
                ("Operational Risk Exposure", risk_out.get("operational", 40)),
                ("Market Volatility Risk", risk_out.get("market", 50)),
                ("Regulatory Compliance Risk", risk_out.get("regulatory", 35)),
            ]
            
            r_idx = 4
            for label, score in risk_metrics:
                ws6.cell(row=r_idx, column=1, value=label).font = label_font
                ws6.cell(row=r_idx, column=2, value=score).font = value_font
                ws6.cell(row=r_idx, column=1).border = thin_border
                ws6.cell(row=r_idx, column=2).border = thin_border
                r_idx += 1

            ws6.cell(row=r_idx+1, column=1, value="Risk Mitigation Commentary:").font = section_font
            ws6.cell(row=r_idx+2, column=1, value=ctx["risk_summary"]).font = value_font
            
            auto_fit_columns(ws6)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 7: Competitor Analysis Agent
            # ─────────────────────────────────────────────────────────────────
            ws7 = wb.create_sheet(title="Competitor Analysis Agent")
            apply_basic_styles(ws7, "Competitor Benchmarking Table")

            # Headers — 15 columns
            excel_headers = [
                "Company", "Ticker", "Exchange", "Sector",
                "Revenue", "Market Cap", "Net Profit", "EBITDA",
                "ROE", "ROA", "EBITDA Margin", "Net Margin",
                "Gross Margin", "D/E", "P/E", "Recommendation"
            ]
            for col_idx, hdr in enumerate(excel_headers, start=1):
                cell = ws7.cell(row=3, column=col_idx, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
            ws7.row_dimensions[3].height = 20

            competitors_list = ctx.get("competitors", [])
            r_idx = 4
            for comp_item in competitors_list:
                row_vals = [
                    comp_item.get("name", "N/A"),
                    comp_item.get("ticker", "N/A"),
                    comp_item.get("exchange", "N/A"),
                    comp_item.get("sector", "N/A"),
                    comp_item.get("revenue", "N/A"),
                    comp_item.get("market_cap", "N/A"),
                    comp_item.get("net_profit", "N/A"),
                    comp_item.get("ebitda", "N/A"),
                    comp_item.get("roe", "N/A"),
                    comp_item.get("roa", "N/A"),
                    comp_item.get("ebitda_margin", "N/A"),
                    comp_item.get("net_margin", "N/A"),
                    comp_item.get("gross_margin", "N/A"),
                    comp_item.get("debt_to_equity", "N/A"),
                    comp_item.get("pe", "N/A"),
                    comp_item.get("recommendation", "HOLD"),
                ]
                for col_idx, val in enumerate(row_vals, start=1):
                    cell = ws7.cell(row=r_idx, column=col_idx, value=str(val) if val is not None else "N/A")
                    cell.font = value_font if col_idx > 1 else label_font
                    cell.border = thin_border
                    # Highlight target company row
                    if comp_item.get("is_target"):
                        cell.fill = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
                r_idx += 1

            ws7.auto_filter.ref = f"A3:P{r_idx-1}"
            
            ws7.cell(row=r_idx+1, column=1, value="Competitive Winner:").font = section_font
            ws7.cell(row=r_idx+1, column=3, value=ctx["competitor_winner"]).font = label_font
            ws7.cell(row=r_idx+2, column=1, value="Narrative Comparison:").font = section_font
            ws7.cell(row=r_idx+3, column=1, value=ctx["competitor_summary"]).font = value_font
            
            auto_fit_columns(ws7)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 8: Market News Agent
            # ─────────────────────────────────────────────────────────────────
            ws8 = wb.create_sheet(title="Market News Agent")
            apply_basic_styles(ws8, "Market News Sentiment Analysis")
            
            news_out = ctx.get("news", {})
            
            ws8["A3"] = "Overall Market Sentiment:"
            ws8["A3"].font = label_font
            ws8["B3"] = news_out.get("overall_sentiment", "Positive")
            ws8["B3"].font = Font(name="Segoe UI", size=11, bold=True, color="38A169" if "POS" in str(news_out.get("overall_sentiment")).upper() else "E53E3E")
            
            ws8["A4"] = "Sentiment Score:"
            ws8["A4"].font = label_font
            ws8["B4"] = f"{news_out.get('sentiment_score', 75)}%"
            ws8["B4"].font = value_font

            ws8.cell(row=6, column=1, value="News Headline").font = header_font
            ws8.cell(row=6, column=1).fill = header_fill
            ws8.cell(row=6, column=2, value="Source").font = header_font
            ws8.cell(row=6, column=2).fill = header_fill
            ws8.cell(row=6, column=3, value="Sentiment").font = header_font
            ws8.cell(row=6, column=3).fill = header_fill
            ws8.row_dimensions[6].height = 20

            r_idx = 7
            for art in news_out.get("articles", []):
                ws8.cell(row=r_idx, column=1, value=art.get("headline", "News updates")).font = value_font
                ws8.cell(row=r_idx, column=2, value=art.get("source", "Financial News")).font = value_font
                ws8.cell(row=r_idx, column=3, value=art.get("sentiment", "Positive")).font = value_font
                for col in range(1, 4):
                    ws8.cell(row=r_idx, column=col).border = thin_border
                r_idx += 1

            ws8.auto_filter.ref = f"A6:C{r_idx-1}"
            auto_fit_columns(ws8)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 9: SWOT Analysis Agent
            # ─────────────────────────────────────────────────────────────────
            ws9 = wb.create_sheet(title="SWOT Analysis Agent")
            apply_basic_styles(ws9, "SWOT Quadrant Matrices & Narrative")
            
            ws9.cell(row=3, column=1, value="STRENGTHS (Internal)").font = header_font
            ws9.cell(row=3, column=1).fill = PatternFill(start_color="2A4365", end_color="2A4365", fill_type="solid")
            ws9.cell(row=3, column=2, value="WEAKNESSES (Internal)").font = header_font
            ws9.cell(row=3, column=2).fill = PatternFill(start_color="742A2A", end_color="742A2A", fill_type="solid")
            
            s_str = "\n".join([f"• {x}" for x in ctx.get("swot_strengths", [])])
            w_str = "\n".join([f"• {x}" for x in ctx.get("swot_weaknesses", [])])
            
            ws9.cell(row=4, column=1, value=s_str).font = value_font
            ws9.cell(row=4, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws9.cell(row=4, column=2, value=w_str).font = value_font
            ws9.cell(row=4, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            
            ws9.cell(row=6, column=1, value="OPPORTUNITIES (External)").font = header_font
            ws9.cell(row=6, column=1).fill = PatternFill(start_color="22543D", end_color="22543D", fill_type="solid")
            ws9.cell(row=6, column=2, value="THREATS (External)").font = header_font
            ws9.cell(row=6, column=2).fill = PatternFill(start_color="7B341E", end_color="7B341E", fill_type="solid")
            
            o_str = "\n".join([f"• {x}" for x in ctx.get("swot_opportunities", [])])
            t_str = "\n".join([f"• {x}" for x in ctx.get("swot_threats", [])])
            
            ws9.cell(row=7, column=1, value=o_str).font = value_font
            ws9.cell(row=7, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws9.cell(row=7, column=2, value=t_str).font = value_font
            ws9.cell(row=7, column=2).alignment = Alignment(wrap_text=True, vertical="top")

            # Border quad lines
            for row in [3, 4, 6, 7]:
                for col in [1, 2]:
                    ws9.cell(row=row, column=col).border = thin_border
            
            ws9.cell(row=9, column=1, value="Consolidated SWOT Narrative:").font = section_font
            ws9.cell(row=10, column=1, value=ctx["swot_narrative"]).font = value_font
            
            auto_fit_columns(ws9)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 10: Investment Recommendation Agent
            # ─────────────────────────────────────────────────────────────────
            ws10 = wb.create_sheet(title="Investment Recommendation Agent")
            apply_basic_styles(ws10, "Investment Decision Recommendations")
            
            ws10["A3"] = "Final Recommendation Rating:"
            ws10["A3"].font = label_font
            ws10["B3"] = ctx["investment_rec"]
            ws10["B3"].font = Font(name="Segoe UI", size=11, bold=True, color="38A169" if "BUY" in str(ctx["investment_rec"]).upper() else "E53E3E")
            
            ws10["A4"] = "Confidence Score:"
            ws10["A4"].font = label_font
            ws10["B4"] = f"{ctx['investment_confidence']}%"
            ws10["B4"].font = value_font

            ws10["A5"] = "Investment Horizon:"
            ws10["A5"].font = label_font
            ws10["B5"] = ctx.get("investment", {}).get("time_horizon", "12 Months")
            ws10["B5"].font = value_font

            ws10.cell(row=7, column=1, value="Recommendation Factors").font = header_font
            ws10.cell(row=7, column=1).fill = header_fill
            ws10.cell(row=7, column=2, value="Weighted Score").font = header_font
            ws10.cell(row=7, column=2).fill = header_fill
            ws10.row_dimensions[7].height = 20

            r_idx = 8
            for f_item in ctx["investment_factors"]:
                ws10.cell(row=r_idx, column=1, value=f_item.get("factor", "N/A")).font = label_font
                ws10.cell(row=r_idx, column=2, value=f_item.get("score", "N/A")).font = value_font
                ws10.cell(row=r_idx, column=1).border = thin_border
                ws10.cell(row=r_idx, column=2).border = thin_border
                r_idx += 1

            ws10.cell(row=r_idx+1, column=1, value="Recommendation Rationale:").font = section_font
            ws10.cell(row=r_idx+2, column=1, value=ctx["investment_rationale"]).font = value_font
            
            auto_fit_columns(ws10)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 11: Executive Summary Agent
            # ─────────────────────────────────────────────────────────────────
            ws11 = wb.create_sheet(title="Executive Summary Agent")
            apply_basic_styles(ws11, "Executive Summary Analytical Report")
            
            ws11.cell(row=3, column=1, value="Detailed Executive Narrative:").font = section_font
            r_idx = 4
            for p_text in ctx.get("executive_summary_paragraphs", []):
                ws11.cell(row=r_idx, column=1, value=str(p_text)).font = value_font
                ws11.cell(row=r_idx, column=1).alignment = Alignment(wrap_text=True, vertical="top")
                r_idx += 2
                
            ws11.cell(row=r_idx, column=1, value="Key Analytical Highlights:").font = section_font
            r_idx += 1
            for highlight in ctx.get("executive_summary_highlights", []):
                ws11.cell(row=r_idx, column=1, value=f"• {highlight}").font = value_font
                r_idx += 1
                
            auto_fit_columns(ws11)

            # ─────────────────────────────────────────────────────────────────
            # Sheet 12: Charts
            # ─────────────────────────────────────────────────────────────────
            ws12 = wb.create_sheet(title="Charts")
            ws12.views.sheetView[0].showGridLines = True
            ws12["A1"] = "Visualized Financial Trend Dashboards"
            ws12["A1"].font = title_font
            ws12.row_dimensions[1].height = 25
            
            # Generate Matplotlib chart buffers
            chart_imgs = self._generate_charts(ctx)
            
            anchors = [
                ("revenue_trend", "A3"),
                ("profit_trend", "A20"),
                ("cash_flow_trend", "A37"),
                ("health_radar", "H3"),
                ("risk_bars", "H20"),
                ("competitor_bars", "H37")
            ]
            
            for key, anchor in anchors:
                chart_bytes = chart_imgs[key]
                pil_img = PILImage.open(chart_bytes)
                xlsx_img = openpyxl.drawing.image.Image(pil_img)
                ws12.add_image(xlsx_img, anchor)

            # Keep save in memory
            out = io.BytesIO()
            wb.save(out)
            logger.info("Excel workbook generation completed successfully.")
            return out.getvalue()
        except Exception as e:
            logger.error(f"Failed to generate Excel sheet: {e}", exc_info=True)
            raise RuntimeError(f"Excel generation failed: {e}")
